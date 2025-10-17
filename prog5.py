# Cusro R Básico
# Python – Estruturas de Controle  
# Instrutor: Lázaro Aparecido
# Instituição INPE
# Data: 17/10/25

## Lendo planilhas

## Mostrando as folhas de uma planilha 

import openpyxl 

wb = openpyxl.load_workbook('medidas_satelite.xlsx') 
print(wb.sheetnames) 

wb.close()

# selecionando uma folha de uma planilha

import openpyxl

wb = openpyxl.load_workbook('medidas_satelite.xlsx') 
folha_atual = wb['modelo_engenharia']

print(folha_atual)

# podemos selecionar uma folha desta outra forma:

folha_atual = wb[wb.sheetnames[0]] 
print(folha_atual)

# ou desta forma tambem:

print(folha_atual.title) 
wb.close()

wb = openpyxl.load_workbook('medidas_satelite.xlsx') 
 
folha_atual = wb['modelo_engenharia'] 
 
var1 = folha_atual['A1'] 
print(var1.value)

# e outra forma 
 
var2 = folha_atual.cell(row = 2, column = 2) 
print(var2.value) 
 
# e para mostrar o numero total de linhas e colunas: 
 
print(folha_atual.max_row) 
print(folha_atual.max_column) 
 
wb.close() 
